import openpyxl
import re
import io
import pandas as pd
from datetime import datetime

from openpyxl import load_workbook, Workbook

# ==============================================================================
# CÁC HÀM TIỆN ÍCH (KHÔNG THAY ĐỔI)
# ==============================================================================

def clean_string(s):
    """Hàm hỗ trợ làm sạch chuỗi, loại bỏ khoảng trắng thừa và dấu nháy đơn ở đầu."""
    if s is None:
        return ""
    cleaned_s = str(s).strip()
    if cleaned_s.startswith("'"):
        cleaned_s = cleaned_s[1:]
    return re.sub(r'\s+', ' ', cleaned_s)

def to_float(value):
    """Hàm hỗ trợ chuyển đổi một giá trị (có thể là text) sang dạng số."""
    if value is None:
        return 0.0
    try:
        return float(str(value).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0

def format_tax_code(raw_vat_value):
    """Chuẩn hóa giá trị VAT sang định dạng chuỗi 2 chữ số (ví dụ: "08", "10")."""
    if raw_vat_value is None:
        return ""
    try:
        s_value = str(raw_vat_value).replace('%', '').strip()
        f_value = float(s_value)
        if 0 < f_value < 1:
            f_value *= 100
        return f"{round(f_value):02d}"
    except (ValueError, TypeError):
        return ""

# ==============================================================================
# CÁC HÀM NẠP DỮ LIỆU TĨNH (KHÔNG THAY ĐỔI)
# ==============================================================================

def load_static_data(data_file_path, mahh_file_path, dskh_file_path):
    """
    Hàm này đọc file Data.xlsx, MaHH.xlsx, và DSKH.xlsx, trả về một dictionary chứa tất cả dữ liệu cấu hình.
    """
    static_data = {}
    try:
        # --- Đọc file Data.xlsx ---
        wb = load_workbook(data_file_path, data_only=True)
        ws = wb.active
        chxd_list, tk_mk_map, khhd_map, chxd_to_khuvuc_map = [], {}, {}, {}
        
        vu_viec_map = {}
        vu_viec_headers = [clean_string(cell.value) for cell in ws[2][4:9]]

        for row_values in ws.iter_rows(min_row=3, max_col=12, values_only=True):
            chxd_name = clean_string(row_values[3])
            if chxd_name:
                ma_kho, khhd, khu_vuc = clean_string(row_values[9]), clean_string(row_values[10]), clean_string(row_values[11])
                if chxd_name not in tk_mk_map: chxd_list.append(chxd_name)
                if ma_kho: tk_mk_map[chxd_name] = ma_kho
                if khhd: khhd_map[chxd_name] = khhd
                if khu_vuc: chxd_to_khuvuc_map[chxd_name] = khu_vuc

                vu_viec_map[chxd_name] = {}
                vu_viec_data_row = row_values[4:9]
                for i, header in enumerate(vu_viec_headers):
                    if header:
                        key = "Dầu mỡ nhờn" if i == len(vu_viec_headers) - 1 else header
                        vu_viec_map[chxd_name][key] = clean_string(vu_viec_data_row[i])

        if not chxd_list: return None, "Không tìm thấy Tên CHXD nào trong cột D của file Data.xlsx."
        static_data.update({
            "DS_CHXD": chxd_list, "tk_mk": tk_mk_map, "khhd_map": khhd_map, 
            "chxd_to_khuvuc_map": chxd_to_khuvuc_map, "vu_viec_map": vu_viec_map
        })

        def get_lookup_map(min_r, max_r, min_c=1, max_c=2):
            return {clean_string(row[0]): row[1] for row in ws.iter_rows(min_row=min_r, max_row=max_r, min_col=min_c, max_col=max_c, values_only=True) if row[0] and row[1] is not None}

        # --- ĐỌC CÁC BẢN ĐỒ TRA CỨU TÀI KHOẢN ---
        phi_bvmt_map_raw = get_lookup_map(10, 13)
        static_data["phi_bvmt_map"] = {k: to_float(v) for k, v in phi_bvmt_map_raw.items()}
        static_data["tk_no_map"] = get_lookup_map(29, 31)
        static_data["tk_doanh_thu_map"] = get_lookup_map(33, 35)
        static_data["tk_thue_co_map"] = get_lookup_map(38, 40)
        static_data["tk_gia_von_value"] = ws['B36'].value
        static_data["tk_no_bvmt_map"] = get_lookup_map(44, 46)
        static_data["tk_dt_thue_bvmt_map"] = get_lookup_map(48, 50)
        static_data["tk_gia_von_bvmt_value"] = ws['B51'].value
        static_data["tk_thue_co_bvmt_map"] = get_lookup_map(53, 55)

        # --- ĐỌC FILE MaHH.xlsx ---
        ma_hang_map = {}
        wb_mahh = load_workbook(mahh_file_path, data_only=True)
        ws_mahh = wb_mahh.active
        for row in ws_mahh.iter_rows(min_row=2, max_col=3, values_only=True):
            ten_hang, ma_hang = clean_string(row[0]), clean_string(row[2])
            if ten_hang and ma_hang:
                ma_hang_map[ten_hang] = ma_hang
        static_data["ma_hang_map"] = ma_hang_map

        # --- ĐỌC FILE DSKH.xlsx ---
        mst_to_makh_map = {}
        wb_dskh = load_workbook(dskh_file_path, data_only=True)
        ws_dskh = wb_dskh.active
        for row in ws_dskh.iter_rows(min_row=2, max_col=4, values_only=True):
            mst = clean_string(row[2])
            ma_kh = clean_string(row[3])
            if mst:
                mst_to_makh_map[mst] = ma_kh
        static_data["mst_to_makh_map"] = mst_to_makh_map
        
        return static_data, None

    except FileNotFoundError as e:
        return None, f"Lỗi: Không tìm thấy file cấu hình. Chi tiết: {e.filename}"
    except Exception as e:
        return None, f"Lỗi khi đọc file cấu hình: {e}"

# ==============================================================================
# CÁC HÀM LOGIC CỐT LÕI (KHÔNG THAY ĐỔI)
# ==============================================================================

def _create_upsse_workbook():
    """Tạo một file Excel mới với cấu trúc tiêu đề cho UpSSE."""
    headers = ["Mã khách", "Tên khách hàng", "Ngày", "Số hóa đơn", "Ký hiệu", "Diễn giải", "Mã hàng", "Tên mặt hàng", "Đvt", "Mã kho", "Mã vị trí", "Mã lô", "Số lượng", "Giá bán", "Tiền hàng", "Mã nt", "Tỷ giá", "Mã thuế", "Tk nợ", "Tk doanh thu", "Tk giá vốn", "Tk thuế có", "Cục thuế", "Vụ việc", "Bộ phận", "Lsx", "Sản phẩm", "Hợp đồng", "Phí", "Khế ước", "Nhân viên bán", "Tên KH(thuế)", "Địa chỉ (thuế)", "Mã số Thuế", "Nhóm Hàng", "Ghi chú", "Tiền thuế"]
    wb = Workbook()
    ws = wb.active
    for _ in range(4): ws.append([''] * len(headers))
    ws.append(headers)
    return wb

def _create_bvmt_row(original_row, phi_bvmt, static_data, khu_vuc):
    """Tạo dòng thuế BVMT dựa trên dòng hóa đơn gốc."""
    bvmt_row = list(original_row)
    tk_no_bvmt = static_data.get('tk_no_bvmt_map', {}).get(khu_vuc)
    tk_dt_thue_bvmt = static_data.get('tk_dt_thue_bvmt_map', {}).get(khu_vuc)
    tk_gia_von_bvmt = static_data.get('tk_gia_von_bvmt_value')
    tk_thue_co_bvmt = static_data.get('tk_thue_co_bvmt_map', {}).get(khu_vuc)
    so_luong = to_float(original_row[12])
    ma_thue = original_row[17]
    thue_suat = to_float(ma_thue) / 100.0 if ma_thue else 0.0
    bvmt_row[6] = "TMT"
    bvmt_row[7] = "Thuế bảo vệ môi trường"
    bvmt_row[13] = phi_bvmt
    bvmt_row[14] = round(phi_bvmt * so_luong)
    bvmt_row[18] = tk_no_bvmt
    bvmt_row[19] = tk_dt_thue_bvmt
    bvmt_row[20] = tk_gia_von_bvmt
    bvmt_row[21] = tk_thue_co_bvmt
    bvmt_row[36] = round(phi_bvmt * so_luong * thue_suat)
    for i in [5, 31, 32, 33]: bvmt_row[i] = ''
    return bvmt_row

def _generate_upsse_from_rows(rows_to_process, static_data, selected_chxd, final_date, summary_suffix_map):
    """
    Hàm lõi: Xử lý một danh sách các dòng từ bảng kê và tạo ra file UpSSE.
    Hàm này được gọi cho mỗi giai đoạn giá.
    """
    if not rows_to_process:
        return None # Trả về None nếu không có dòng nào để xử lý

    # --- Lấy các dữ liệu cấu hình cần thiết ---
    khu_vuc = static_data['chxd_to_khuvuc_map'].get(selected_chxd)
    ma_kho = static_data['tk_mk'].get(selected_chxd)
    ma_hang_map = static_data['ma_hang_map']
    phi_bvmt_map = static_data['phi_bvmt_map']
    vu_viec_map = static_data['vu_viec_map']
    mst_to_makh_map = static_data['mst_to_makh_map']
    xang_dau_group = ["Xăng E5 RON 92-II", "Xăng RON 95-III", "Dầu DO 0,05S-II", "Dầu DO 0,001S-V"]
    
    tk_no = static_data['tk_no_map'].get(khu_vuc)
    tk_doanh_thu = static_data['tk_doanh_thu_map'].get(khu_vuc)
    tk_gia_von = static_data['tk_gia_von_value']
    tk_thue_co = static_data['tk_thue_co_map'].get(khu_vuc)

    # --- Bắt đầu xử lý ---
    original_invoice_rows = []
    bvmt_rows = []
    summary_data = {}
    first_invoice_prefix_source = ""

    for bkhd_row in rows_to_process:
        if to_float(bkhd_row[8] if len(bkhd_row) > 8 else None) <= 0: continue

        ten_kh = clean_string(bkhd_row[3])
        ten_mat_hang = clean_string(bkhd_row[6])
        is_anonymous = (ten_kh == "Người mua không lấy hóa đơn")
        is_petrol_product = (ten_mat_hang in xang_dau_group)
        
        # Xử lý hóa đơn riêng lẻ (không phải khách vãng lai mua xăng dầu)
        if not is_anonymous or not is_petrol_product:
            new_upsse_row = [''] * 37
            new_upsse_row[9] = ma_kho
            new_upsse_row[1] = ten_kh
            new_upsse_row[31] = ten_kh
            new_upsse_row[2] = final_date
            ky_hieu_shd = str(bkhd_row[18] or '').strip()
            so_hd_goc = str(bkhd_row[19] or '').strip()
            so_hoa_don_moi = f"HN{so_hd_goc[-6:]}" if selected_chxd == "Nguyễn Huệ" else f"{ky_hieu_shd[-2:]}{so_hd_goc[-6:]}"
            new_upsse_row[3] = so_hoa_don_moi
            new_upsse_row[4] = clean_string(bkhd_row[17]) + clean_string(bkhd_row[18])
            new_upsse_row[5] = "Xuất bán hàng theo hóa đơn số " + so_hoa_don_moi
            new_upsse_row[7] = ten_mat_hang
            new_upsse_row[6] = ma_hang_map.get(ten_mat_hang, '')
            new_upsse_row[8] = clean_string(bkhd_row[10])
            so_luong = to_float(bkhd_row[8])
            new_upsse_row[12] = round(so_luong, 3)
            don_gia = to_float(bkhd_row[9])
            phi_bvmt = phi_bvmt_map.get(ten_mat_hang, 0.0) if is_petrol_product else 0.0
            gia_ban = don_gia - phi_bvmt
            new_upsse_row[13] = gia_ban
            vat_raw = bkhd_row[14]
            ma_thue = format_tax_code(vat_raw)
            new_upsse_row[17] = ma_thue
            thue_suat = to_float(ma_thue) / 100.0 if ma_thue else 0.0
            tien_thue_goc = to_float(bkhd_row[15])
            tien_thue_phi_bvmt = round(phi_bvmt * so_luong * thue_suat)
            tien_thue_moi = tien_thue_goc - tien_thue_phi_bvmt
            new_upsse_row[36] = round(tien_thue_moi)
            if is_petrol_product:
                phai_thu = to_float(bkhd_row[16])
                tien_hang_phi_bvmt = round(phi_bvmt * so_luong)
                tien_hang = phai_thu - tien_thue_goc - tien_hang_phi_bvmt
            else:
                tien_hang = to_float(bkhd_row[13])
            new_upsse_row[14] = round(tien_hang)
            new_upsse_row[18] = tk_no
            new_upsse_row[19] = tk_doanh_thu
            new_upsse_row[20] = tk_gia_von
            new_upsse_row[21] = tk_thue_co
            chxd_vu_viec_map = vu_viec_map.get(selected_chxd, {})
            ma_vu_viec = chxd_vu_viec_map.get(ten_mat_hang, chxd_vu_viec_map.get("Dầu mỡ nhờn", ''))
            new_upsse_row[23] = ma_vu_viec
            new_upsse_row[32] = clean_string(bkhd_row[4])
            mst_khach_hang = clean_string(bkhd_row[5])
            new_upsse_row[33] = mst_khach_hang
            ma_kh_fast = clean_string(bkhd_row[2])
            ma_khach_final = ma_kho
            if ma_kh_fast and len(ma_kh_fast) < 12:
                ma_khach_final = ma_kh_fast
            elif mst_khach_hang and mst_to_makh_map.get(mst_khach_hang):
                ma_khach_final = mst_to_makh_map.get(mst_khach_hang)
            new_upsse_row[0] = ma_khach_final
            
            original_invoice_rows.append(new_upsse_row)
            if is_petrol_product:
                bvmt_rows.append(_create_bvmt_row(new_upsse_row, phi_bvmt, static_data, khu_vuc))
        
        # Gom dữ liệu khách vãng lai mua xăng dầu
        else:
            if not first_invoice_prefix_source:
                first_invoice_prefix_source = str(bkhd_row[18] or '').strip()
            if ten_mat_hang not in summary_data:
                summary_data[ten_mat_hang] = {
                    'total_so_luong_bkhd': 0, 'total_tien_thue_goc': 0, 'total_phai_thu': 0,
                    'first_invoice_data': {'ky_hieu_mau_so': clean_string(bkhd_row[17]),'ky_hieu_ky_hieu': clean_string(bkhd_row[18]),'don_gia': to_float(bkhd_row[9]),'vat_raw': bkhd_row[14]}
                }
            summary_data[ten_mat_hang]['total_so_luong_bkhd'] += to_float(bkhd_row[8])
            summary_data[ten_mat_hang]['total_tien_thue_goc'] += to_float(bkhd_row[15])
            summary_data[ten_mat_hang]['total_phai_thu'] += to_float(bkhd_row[16])

    # --- Tạo các dòng tổng hợp cho khách vãng lai ---
    prefix = first_invoice_prefix_source[-2:] if len(first_invoice_prefix_source) >= 2 else first_invoice_prefix_source
    for product_name, data in summary_data.items():
        summary_row = [''] * 37
        first_data = data['first_invoice_data']
        date_part = f"{final_date.day:02d}.{final_date.month:02d}"
        suffix = summary_suffix_map.get(product_name, "")
        summary_invoice_number = f"{prefix}BK.{date_part}.{suffix}"
        total_so_luong = data['total_so_luong_bkhd']
        phi_bvmt_unit = phi_bvmt_map.get(product_name, 0.0)
        ma_thue = format_tax_code(first_data['vat_raw'])
        thue_suat = to_float(ma_thue) / 100.0 if ma_thue else 0.0
        TDT, TTT = data['total_phai_thu'], data['total_tien_thue_goc']
        TH_TMT = round(phi_bvmt_unit * total_so_luong)
        TT_TMT = round(TH_TMT * thue_suat)
        TT_goc = TTT - TT_TMT
        TH_goc = TDT - TH_TMT - TT_goc - TT_TMT
        summary_row[0] = ma_kho
        summary_row[1] = f"Khách hàng mua {product_name} không lấy hóa đơn"
        summary_row[31] = summary_row[1]
        summary_row[2] = final_date
        summary_row[3] = summary_invoice_number
        summary_row[4] = first_data['ky_hieu_mau_so'] + first_data['ky_hieu_ky_hieu']
        summary_row[5] = "Xuất bán hàng theo hóa đơn số " + summary_invoice_number
        summary_row[7] = product_name
        summary_row[6] = ma_hang_map.get(product_name, '')
        summary_row[8] = "Lít"
        summary_row[9] = ma_kho
        summary_row[12] = round(total_so_luong, 3)
        summary_row[13] = first_data['don_gia'] - phi_bvmt_unit
        summary_row[14] = round(TH_goc)
        summary_row[17] = ma_thue
        summary_row[18] = tk_no
        summary_row[19] = tk_doanh_thu
        summary_row[20] = tk_gia_von
        summary_row[21] = tk_thue_co
        chxd_vu_viec_map = vu_viec_map.get(selected_chxd, {})
        summary_row[23] = chxd_vu_viec_map.get(product_name, '')
        summary_row[36] = round(TT_goc)
        original_invoice_rows.append(summary_row)
        bvmt_rows.append(_create_bvmt_row(summary_row, phi_bvmt_unit, static_data, khu_vuc))

    # --- Ghi ra file Excel trong bộ nhớ ---
    upsse_wb = _create_upsse_workbook()
    upsse_ws = upsse_wb.active
    all_final_rows = original_invoice_rows + bvmt_rows
    for row_data in all_final_rows:
        upsse_ws.append(row_data)
    for row_index in range(6, upsse_ws.max_row + 1):
        upsse_ws[f'C{row_index}'].number_format = 'dd/mm/yyyy'
    output_buffer = io.BytesIO()
    upsse_wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer

# ==============================================================================
# HÀM CHÍNH ĐIỀU PHỐI (MAIN DISPATCHER)
# ==============================================================================

def _load_uploaded_workbook(file_content_bytes):
    """Đọc file người dùng tải lên trực tiếp như một file Excel."""
    try:
        return load_workbook(io.BytesIO(file_content_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Lỗi khi đọc file Bảng kê hóa đơn. Hãy chắc chắn file tải lên là file Excel. Lỗi: {e}")

def _analyze_date_ambiguity(worksheet):
    """Phân tích ngày tháng trong BKHD."""
    unique_dates = set()
    for row in worksheet.iter_rows(min_row=11, values_only=True):
        if to_float(row[8] if len(row) > 8 else None) > 0:
            date_val = row[20] if len(row) > 20 else None
            if isinstance(date_val, datetime):
                unique_dates.add(date_val.date())
            elif isinstance(date_val, (int, float)):
                try:
                    converted_date_obj = pd.to_datetime(date_val, unit='D', origin='1899-12-30').to_pydatetime()
                    unique_dates.add(converted_date_obj.date())
                except (ValueError, TypeError): pass
    if len(unique_dates) > 1: raise ValueError("Công cụ chỉ chạy được khi bạn kết xuất hóa đơn trong 1 ngày duy nhất.")
    if not unique_dates: raise ValueError("Không tìm thấy dữ liệu hóa đơn hợp lệ nào trong file Bảng kê.")
    the_date = unique_dates.pop()
    if the_date.day > 12: return False, datetime(the_date.year, the_date.month, the_date.day), None
    try:
        # [SỬA LỖI] Thay thế a 'the_day' không xác định bằng 'the_date.day'
        date_as_is, swapped_date = datetime(the_date.year, the_date.month, the_date.day), datetime(the_date.year, the_date.day, the_date.month)
        return (date_as_is != swapped_date), date_as_is, swapped_date
    except ValueError:
        return False, datetime(the_date.year, the_date.month, the_date.day), None

def _validate_input(worksheet, selected_chxd, khhd_map):
    """
    Kiểm tra các điều kiện đầu vào.
    [CẢI TIẾN] Quét hết file và báo cáo tất cả các lỗi địa chỉ quá dài cùng một lúc.
    """
    khhd_from_data = khhd_map.get(selected_chxd)
    if not khhd_from_data: raise ValueError(f"Lỗi cấu hình: Không tìm thấy 'Ký hiệu hóa đơn' cho '{selected_chxd}'.")
    khhd_suffix_expected = khhd_from_data[-6:]
    validation_value_raw = worksheet['S11'].value
    if khhd_suffix_expected not in clean_string(validation_value_raw):
        raise ValueError(f"Bảng kê hóa đơn không khớp. Ký hiệu trên file: '{clean_string(validation_value_raw)}', mong đợi chứa: '{khhd_suffix_expected}'.")
    
    # Danh sách để lưu trữ tất cả các lỗi về độ dài địa chỉ
    long_address_errors = []
    
    # Bắt đầu lặp qua các dòng từ dòng 11
    for row_index, row_values in enumerate(worksheet.iter_rows(min_row=11, values_only=True), start=11):
        # Chỉ xử lý những dòng có số lượng > 0 (cột I)
        if to_float(row_values[8] if len(row_values) > 8 else None) > 0:
            # Lấy địa chỉ từ cột E (index 4)
            address = str(row_values[4]) if len(row_values) > 4 and row_values[4] is not None else ""
            
            # Kiểm tra nếu độ dài địa chỉ vượt quá 128 ký tự
            if len(address) > 128:
                # Thêm thông báo lỗi vào danh sách thay vì dừng lại ngay
                long_address_errors.append(f" - Dòng {row_index} (ô E{row_index}): địa chỉ dài {len(address)} ký tự.")

    # Sau khi quét hết file, nếu có lỗi, tạo một thông báo tổng hợp và dừng chương trình
    if long_address_errors:
        error_message = f"Phát hiện {len(long_address_errors)} lỗi địa chỉ quá dài (> 128 ký tự). Vui lòng sửa trong file Excel:\n"
        error_message += "\n".join(long_address_errors)
        raise ValueError(error_message)

def process_uploaded_file(uploaded_file_content, static_data, selected_chxd, price_periods, new_price_invoice_number, confirmed_date_str=None):
    """
    Hàm chính để xử lý file bảng kê.
    Điều phối việc xử lý cho 1 hoặc 2 giai đoạn giá.
    """
    bkhd_wb = _load_uploaded_workbook(uploaded_file_content)
    bkhd_ws = bkhd_wb.active
    
    # --- Giai đoạn 1: Xác định ngày tháng ---
    final_date = None
    if confirmed_date_str:
        final_date = datetime.strptime(confirmed_date_str, '%Y-%m-%d')
    else:
        is_ambiguous, date1, date2 = _analyze_date_ambiguity(bkhd_ws)
        if is_ambiguous:
            return {'choice_needed': True, 'options': [{'text': date1.strftime('%d/%m/%Y'), 'value': date1.strftime('%Y-%m-%d')}, {'text': date2.strftime('%d/%m/%Y'), 'value': date2.strftime('%Y-%m-%d')}]}
        final_date = date1

    # --- Giai đoạn 2: Kiểm tra dữ liệu và điều phối ---
    _validate_input(bkhd_ws, selected_chxd, static_data.get('khhd_map', {}))

    # --- Xử lý cho 1 giai đoạn giá ---
    if price_periods == '1':
        all_rows = list(bkhd_ws.iter_rows(min_row=11, values_only=True))
        suffix_map = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}
        return _generate_upsse_from_rows(all_rows, static_data, selected_chxd, final_date, suffix_map)
    
    # --- Xử lý cho 2 giai đoạn giá ---
    elif price_periods == '2':
        if not new_price_invoice_number:
            raise ValueError("Vui lòng nhập 'Số hóa đơn đầu tiên của giá mới' khi chọn 2 giai đoạn giá.")

        all_rows = list(bkhd_ws.iter_rows(min_row=11, values_only=True))
        split_index = -1
        for i, row in enumerate(all_rows):
            invoice_num_in_row = str(row[19] or '').strip() # Cột T là số hóa đơn
            if invoice_num_in_row == new_price_invoice_number:
                split_index = i
                break
        
        if split_index == -1:
            raise ValueError(f"Không tìm thấy hóa đơn có số '{new_price_invoice_number}' trong file Bảng kê. Vui lòng kiểm tra lại.")

        old_price_rows = all_rows[:split_index]
        new_price_rows = all_rows[split_index:]

        if not old_price_rows:
            raise ValueError("Không có dữ liệu cho giai đoạn giá cũ (trước hóa đơn đã nhập). Vui lòng kiểm tra lại số hóa đơn.")
        
        suffix_map_old = {"Xăng E5 RON 92-II": "1", "Xăng RON 95-III": "2", "Dầu DO 0,05S-II": "3", "Dầu DO 0,001S-V": "4"}
        suffix_map_new = {"Xăng E5 RON 92-II": "5", "Xăng RON 95-III": "6", "Dầu DO 0,05S-II": "7", "Dầu DO 0,001S-V": "8"}

        result_old = _generate_upsse_from_rows(old_price_rows, static_data, selected_chxd, final_date, suffix_map_old)
        result_new = _generate_upsse_from_rows(new_price_rows, static_data, selected_chxd, final_date, suffix_map_new)

        # Nếu một trong hai giai đoạn không có dữ liệu để xử lý (ví dụ, toàn bộ là hóa đơn có giá trị 0)
        if not result_old and not result_new:
             raise ValueError("Không có dữ liệu hóa đơn hợp lệ trong cả hai giai đoạn giá.")
        if not result_old:
             return result_new # Chỉ trả về file giá mới nếu giá cũ không có gì
        if not result_new:
             return result_old # Chỉ trả về file giá cũ nếu giá mới không có gì

        return {'old': result_old, 'new': result_new}
    
    else:
        raise ValueError("Lựa chọn giai đoạn giá không hợp lệ.")
